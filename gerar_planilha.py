# -*- coding: utf-8 -*-
"""
Gerador da Planilha Financeira Pessoal
Tema: preto, verde e branco | Compatível com Excel e Google Planilhas
Execute:  python3 gerar_planilha.py  ->  Planilha_Financeira_Pessoal.xlsx
"""
import datetime as dt

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

# ----------------------------------------------------------------------------
# Paleta (preto / verde / branco)
# ----------------------------------------------------------------------------
BG        = "0B0F0C"   # fundo geral (quase preto)
CARD      = "141B16"   # cartões / tabelas
HEADER    = "0F2418"   # cabeçalhos de tabela
GREEN     = "00C853"   # verde principal
GREEN_HI  = "69F0AE"   # verde claro (destaques)
GREEN_DK  = "1B5E20"   # verde escuro
WHITE     = "FFFFFF"
GRAY      = "9CB8A6"   # texto secundário
RED       = "FF5252"
AMBER     = "FFD600"

FILL_BG     = PatternFill("solid", fgColor=BG)
FILL_CARD   = PatternFill("solid", fgColor=CARD)
FILL_HEADER = PatternFill("solid", fgColor=HEADER)
FILL_GREEN  = PatternFill("solid", fgColor=GREEN)
FILL_RED    = PatternFill("solid", fgColor="7F1D1D")
FILL_AMBERD = PatternFill("solid", fgColor="713F12")
FILL_GREEND = PatternFill("solid", fgColor="14532D")

F_TITLE  = Font(name="Calibri", size=18, bold=True, color=GREEN_HI)
F_SUB    = Font(name="Calibri", size=11, color=GRAY)
F_HEAD   = Font(name="Calibri", size=11, bold=True, color=GREEN_HI)
F_TEXT   = Font(name="Calibri", size=11, color=WHITE)
F_KPI    = Font(name="Calibri", size=16, bold=True, color=WHITE)
F_KPI_G  = Font(name="Calibri", size=16, bold=True, color=GREEN_HI)
F_LABEL  = Font(name="Calibri", size=10, bold=True, color=GRAY)
F_NAV    = Font(name="Calibri", size=10, bold=True, color=WHITE)
F_SECT   = Font(name="Calibri", size=13, bold=True, color=GREEN)

THIN  = Side(style="thin", color="1F3527")
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ACCENT_BOT = Border(bottom=Side(style="thick", color=GREEN))

C_LEFT   = Alignment(horizontal="left",   vertical="center")
C_CENTER = Alignment(horizontal="center", vertical="center")
C_RIGHT  = Alignment(horizontal="right",  vertical="center")

FMT_MONEY = 'R$ #,##0.00'
FMT_PCT   = '0.0%'
FMT_DATE  = 'DD/MM/YYYY'
FMT_MESANO = 'mmm/yyyy'

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
         "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
MESES_FULL = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
              "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

CAT_REC = ["Salário", "Freelance", "Investimentos", "Vendas", "Aluguel",
           "Reembolso", "Outros"]
CAT_DES = ["Moradia", "Alimentação", "Transporte", "Saúde", "Educação",
           "Lazer", "Vestuário", "Assinaturas", "Contas Fixas",
           "Investimentos", "Outros"]
FORMAS  = ["Dinheiro", "Pix", "Débito", "Crédito", "Boleto", "Transferência"]
CONTAS  = ["Itaú", "Nubank", "Caixa", "Inter", "Carteira"]
CARTOES = ["Nubank", "Itaú", "Inter"]

ANO = 2026
N_ROWS = 200            # linhas de lançamento pré-formatadas
ROW1 = 5                # primeira linha de dados nas abas de lançamentos
ROWN = ROW1 + N_ROWS - 1

SH_DASH = "📊 Dashboard"
SH_REC  = "💰 Receitas"
SH_DES  = "💸 Despesas"
SH_CON  = "🏦 Contas"
SH_CAR  = "💳 Cartões"
SH_MET  = "🎯 Metas"
SH_PLA  = "📅 Planejamento"
SH_REL  = "📈 Relatórios"
SH_ALE  = "🔔 Alertas"
SH_CFG  = "⚙️ Config"
ALL_SHEETS = [SH_DASH, SH_REC, SH_DES, SH_CON, SH_CAR,
              SH_MET, SH_PLA, SH_REL, SH_ALE, SH_CFG]
NAV_LABELS = ["📊 Início", "💰 Receitas", "💸 Despesas", "🏦 Contas",
              "💳 Cartões", "🎯 Metas", "📅 Orçamento", "📈 Relatórios",
              "🔔 Alertas", "⚙️ Config"]

Q = lambda s: f"'{s}'"   # aspas para nomes de aba nas fórmulas

MES_CELL = f"{Q(SH_DASH)}!$J$3"
ANO_CELL = f"{Q(SH_DASH)}!$L$3"
INI_MES  = f"DATE({ANO_CELL},{MES_CELL},1)"
FIM_MES  = f"EOMONTH(DATE({ANO_CELL},{MES_CELL},1),0)"


# ----------------------------------------------------------------------------
# Auxiliares
# ----------------------------------------------------------------------------
def put(ws, row, col, value=None, font=F_TEXT, fill=FILL_CARD, fmt=None,
        align=C_LEFT, border=BORD, unlock=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = font
    c.fill = fill
    c.alignment = align
    if border is not None:
        c.border = border
    if fmt:
        c.number_format = fmt
    if unlock:
        c.protection = Protection(locked=False)
    return c


def paint_bg(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_BG
            cell.font = F_TEXT


def nav_bar(ws, current):
    ws.row_dimensions[1].height = 24
    for i, (name, label) in enumerate(zip(ALL_SHEETS, NAV_LABELS)):
        col = 2 + i
        c = ws.cell(row=1, column=col, value=label)
        c.font = Font(name="Calibri", size=10, bold=True,
                      color="0B0F0C" if name == current else WHITE)
        c.fill = FILL_GREEN if name == current else PatternFill("solid", fgColor="163020")
        c.alignment = C_CENTER
        c.border = BORD
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{name}'!A1",
                                tooltip=f"Ir para {name}")


def title(ws, row, col, text, subtitle=None):
    put(ws, row, col, text, font=F_TITLE, fill=FILL_BG, border=None)
    if subtitle:
        put(ws, row + 1, col, subtitle, font=F_SUB, fill=FILL_BG, border=None)


def table_header(ws, row, col0, headers):
    for i, h in enumerate(headers):
        put(ws, row, col0 + i, h, font=F_HEAD, fill=FILL_HEADER,
            align=C_CENTER)
        ws.cell(row=row, column=col0 + i).border = Border(
            left=THIN, right=THIN, top=THIN,
            bottom=Side(style="medium", color=GREEN))


def set_widths(ws, widths, start=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start + i)].width = w


def protect(ws):
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def section(ws, row, col, text):
    put(ws, row, col, text, font=F_SECT, fill=FILL_BG, border=None)


# ----------------------------------------------------------------------------
# Workbook
# ----------------------------------------------------------------------------
wb = Workbook()
wb.calculation.fullCalcOnLoad = True

ws_dash = wb.active
ws_dash.title = SH_DASH
ws_rec = wb.create_sheet(SH_REC)
ws_des = wb.create_sheet(SH_DES)
ws_con = wb.create_sheet(SH_CON)
ws_car = wb.create_sheet(SH_CAR)
ws_met = wb.create_sheet(SH_MET)
ws_pla = wb.create_sheet(SH_PLA)
ws_rel = wb.create_sheet(SH_REL)
ws_ale = wb.create_sheet(SH_ALE)
ws_cfg = wb.create_sheet(SH_CFG)

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN_DK
ws_dash.sheet_properties.tabColor = GREEN

# ----------------------------------------------------------------------------
# ⚙️ CONFIG  (listas, nomes de meses, instruções)
# ----------------------------------------------------------------------------
ws = ws_cfg
paint_bg(ws, 40, 14)
nav_bar(ws, SH_CFG)
title(ws, 3, 2, "⚙️ Configurações", "Listas usadas nas validações — edite à vontade.")

cfg_cols = [
    ("Categorias de Receita", CAT_REC),
    ("Categorias de Despesa", CAT_DES),
    ("Formas de Pagamento", FORMAS),
    ("Contas / Bancos", CONTAS),
    ("Cartões", CARTOES),
    ("Mês (nº)", list(range(1, 13))),
    ("Mês (nome)", MESES_FULL),
    ("Anos", list(range(2024, 2032))),
]
for j, (head, values) in enumerate(cfg_cols):
    col = 2 + j
    put(ws, 6, col, head, font=F_HEAD, fill=FILL_HEADER, align=C_CENTER)
    for k in range(18):
        v = values[k] if k < len(values) else None
        put(ws, 7 + k, col, v, unlock=True)
set_widths(ws, [3] + [19] * 9)

put(ws, 27, 2, "💡 Como usar", font=F_SECT, fill=FILL_BG, border=None)
dicas = [
    "1. Lance suas receitas e despesas nas abas 💰 e 💸 (use os menus suspensos).",
    "2. Escolha o mês e o ano no topo do 📊 Dashboard — tudo recalcula sozinho.",
    "3. Cadastre bancos em 🏦 Contas e cartões em 💳 Cartões.",
    "4. Defina orçamentos em 📅 Planejamento e metas em 🎯 Metas.",
    "5. A aba 🔔 Alertas avisa sobre vencimentos, estouros e metas atingidas.",
    "6. As fórmulas estão protegidas (sem senha): Revisão ▸ Desproteger Planilha.",
]
for i, d in enumerate(dicas):
    put(ws, 28 + i, 2, d, font=F_SUB, fill=FILL_BG, border=None)

# Intervalos nomeados p/ validação
def add_name(name, ref):
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName(name, attr_text=f"{Q(SH_CFG)}!{ref}"))

add_name("CatReceitas", "$B$7:$B$18")
add_name("CatDespesas", "$C$7:$C$18")
add_name("FormasPag",   "$D$7:$D$18")
add_name("ListaContas", "$E$7:$E$18")
add_name("ListaCartoes","$F$7:$F$18")
add_name("ListaMeses",  "$G$7:$G$18")
add_name("ListaAnos",   "$I$7:$I$18")

# ----------------------------------------------------------------------------
# 💰 RECEITAS  (A Data | B Descrição | C Categoria | D Valor | E Forma | F Conta | G Status)
# ----------------------------------------------------------------------------
ws = ws_rec
paint_bg(ws, ROWN + 3, 12)
nav_bar(ws, SH_REC)
title(ws, 3, 2, "💰 Controle de Receitas")
put(ws, 3, 6, f'="Total recebido: "&TEXT(SUMIFS($D${ROW1}:$D${ROWN},$G${ROW1}:$G${ROWN},"Recebido")/1000,"0.0")&" mil | Pendente: "&TEXT(SUMIFS($D${ROW1}:$D${ROWN},$G${ROW1}:$G${ROWN},"Pendente")/1000,"0.0")&" mil"',
    font=F_SUB, fill=FILL_BG, border=None)

REC_HEADERS = ["Data", "Descrição", "Categoria", "Valor",
               "Forma de Recebimento", "Conta", "Status"]
table_header(ws, 4, 1, REC_HEADERS)
set_widths(ws, [12, 32, 16, 14, 20, 13, 13])
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:G{ROWN}"

receitas = []
for m in range(1, 8):
    receitas.append((dt.date(ANO, m, 5), "Salário mensal", "Salário", 4800,
                     "Transferência", "Itaú", "Recebido"))
freelas = [(1, 18, "Projeto de site — freelance", 800, "Recebido"),
           (3, 22, "Consultoria de marketing", 1200, "Recebido"),
           (5, 15, "Criação de logotipo", 950, "Recebido"),
           (7, 20, "Manutenção de site", 600, "Pendente")]
for m, d, desc, v, st in freelas:
    receitas.append((dt.date(ANO, m, d), desc, "Freelance", v, "Pix", "Nubank", st))
for i, (m, v) in enumerate(zip(range(2, 8), [130, 145, 150, 138, 160, 155])):
    receitas.append((dt.date(ANO, m, 10), "Dividendos de FIIs", "Investimentos",
                     v, "Transferência", "Inter", "Recebido"))
receitas.sort(key=lambda r: r[0])

for r in range(ROW1, ROWN + 1):
    data = receitas[r - ROW1] if r - ROW1 < len(receitas) else [None] * 7
    fmts = [FMT_DATE, None, None, FMT_MONEY, None, None, None]
    aligns = [C_CENTER, C_LEFT, C_LEFT, C_RIGHT, C_LEFT, C_LEFT, C_CENTER]
    for j in range(7):
        put(ws, r, 1 + j, data[j], fmt=fmts[j], align=aligns[j], unlock=True)

dv = DataValidation(type="list", formula1="=CatReceitas", allow_blank=True)
ws.add_data_validation(dv); dv.add(f"C{ROW1}:C{ROWN}")
dv = DataValidation(type="list", formula1="=FormasPag", allow_blank=True)
ws.add_data_validation(dv); dv.add(f"E{ROW1}:E{ROWN}")
dv = DataValidation(type="list", formula1="=ListaContas", allow_blank=True)
ws.add_data_validation(dv); dv.add(f"F{ROW1}:F{ROWN}")
dv = DataValidation(type="list", formula1='"Recebido,Pendente"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"G{ROW1}:G{ROWN}")

rng = f"G{ROW1}:G{ROWN}"
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'$G{ROW1}="Recebido"'], font=Font(color=GREEN_HI, bold=True)))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'$G{ROW1}="Pendente"'], font=Font(color=AMBER, bold=True)))

# ----------------------------------------------------------------------------
# 💸 DESPESAS (A Data | B Descrição | C Categoria | D Valor | E Forma | F Cartão
#              | G Parcelas | H Vencimento | I Pago | J Conta | K rank-alerta)
# ----------------------------------------------------------------------------
ws = ws_des
paint_bg(ws, ROWN + 3, 13)
nav_bar(ws, SH_DES)
title(ws, 3, 2, "💸 Controle de Despesas")
put(ws, 3, 6, f'="Em aberto: "&COUNTIFS($I${ROW1}:$I${ROWN},"Não")&" conta(s) não paga(s)"',
    font=F_SUB, fill=FILL_BG, border=None)

DES_HEADERS = ["Data", "Descrição", "Categoria", "Valor", "Forma de Pagamento",
               "Cartão", "Parcelas", "Vencimento", "Pago", "Conta", "•"]
table_header(ws, 4, 1, DES_HEADERS)
set_widths(ws, [12, 32, 15, 13, 18, 11, 10, 13, 9, 11, 5])
ws.column_dimensions["K"].hidden = True
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:J{ROWN}"

def venc_fatura(m, dia=5):
    """vencimento da fatura no mês seguinte"""
    return dt.date(ANO if m < 12 else ANO + 1, m % 12 + 1, dia)

despesas = []
mercado = [620, 580, 710, 650, 690, 720, 680]
energia = [175, 182, 160, 150, 140, 135, 148]
combust = [240, 220, 260, 230, 250, 270, 90]
restaur = [180, 210, 150, 240, 190, 260, 120]
for m in range(1, 8):
    pago = "Sim" if m < 7 else "Não"
    despesas += [
        (dt.date(ANO, m, 8), "Aluguel + condomínio", "Moradia", 1500, "Boleto",
         "", "", dt.date(ANO, m, 10), pago, "Itaú"),
        (dt.date(ANO, m, 12), "Supermercado do mês", "Alimentação", mercado[m-1],
         "Crédito", "Nubank", "", venc_fatura(m), pago, "Nubank"),
        (dt.date(ANO, m, 3), "Energia elétrica", "Contas Fixas", energia[m-1],
         "Pix", "", "", dt.date(ANO, m, 8), pago, "Itaú"),
        (dt.date(ANO, m, 13), "Internet fibra", "Contas Fixas", 100, "Débito",
         "", "", dt.date(ANO, m, 15), pago, "Itaú"),
        (dt.date(ANO, m, 10), "Plano de celular", "Contas Fixas", 60, "Débito",
         "", "", dt.date(ANO, m, 12), pago, "Nubank"),
        (dt.date(ANO, m, 4), "Academia", "Saúde", 90, "Débito", "", "",
         dt.date(ANO, m, 5), pago, "Nubank"),
        (dt.date(ANO, m, 6), "Streaming (Netflix + Spotify)", "Assinaturas", 55,
         "Crédito", "Nubank", "", venc_fatura(m), pago, "Nubank"),
        (dt.date(ANO, m, 16), "Combustível", "Transporte", combust[m-1],
         "Crédito", "Itaú", "", venc_fatura(m, 3), pago, "Itaú"),
        (dt.date(ANO, m, 22), "Restaurantes e delivery", "Lazer", restaur[m-1],
         "Crédito", "Nubank", "", venc_fatura(m), pago, "Nubank"),
    ]
    if m <= 6:
        despesas.append((dt.date(ANO, m, 6), "Aporte em investimentos",
                         "Investimentos", 500, "Transferência", "", "",
                         dt.date(ANO, m, 6), "Sim", "Itaú"))
for m in range(2, 8):   # notebook 12x — parcelas 1 a 6
    pago = "Sim" if m < 7 else "Não"
    despesas.append((dt.date(ANO, m, 1), f"Notebook Dell — parcela {m-1}/12",
                     "Outros", 300, "Crédito", "Itaú", "12x",
                     venc_fatura(m, 3), pago, "Itaú"))
for m in range(5, 8):   # smartphone 10x — parcelas 1 a 3
    pago = "Sim" if m < 7 else "Não"
    despesas.append((dt.date(ANO, m, 2), f"Smartphone — parcela {m-4}/10",
                     "Outros", 240, "Crédito", "Nubank", "10x",
                     venc_fatura(m), pago, "Nubank"))
despesas += [
    (dt.date(ANO, 1, 20), "Roupas de trabalho", "Vestuário", 180, "Crédito",
     "Itaú", "", venc_fatura(1, 3), "Sim", "Itaú"),
    (dt.date(ANO, 2, 9), "IPVA — parcela única", "Transporte", 310, "Boleto",
     "", "", dt.date(ANO, 2, 15), "Sim", "Itaú"),
    (dt.date(ANO, 3, 14), "Consulta médica", "Saúde", 150, "Pix", "", "",
     dt.date(ANO, 3, 14), "Sim", "Nubank"),
    (dt.date(ANO, 4, 2), "Curso online de inglês", "Educação", 120, "Crédito",
     "Nubank", "", venc_fatura(4), "Sim", "Nubank"),
    (dt.date(ANO, 5, 25), "Tênis de corrida", "Vestuário", 220, "Crédito",
     "Inter", "", venc_fatura(5, 27), "Sim", "Inter"),
    (dt.date(ANO, 6, 18), "Presente de aniversário", "Outros", 130, "Pix",
     "", "", dt.date(ANO, 6, 18), "Sim", "Nubank"),
    (dt.date(ANO, 7, 1), "Cinema com amigos", "Lazer", 65, "Débito", "", "",
     dt.date(ANO, 7, 1), "Sim", "Nubank"),
    (dt.date(ANO, 7, 2), "Ingresso de show", "Lazer", 120, "Crédito", "Nubank",
     "", venc_fatura(7), "Não", "Nubank"),
]
despesas.sort(key=lambda r: r[0])

for r in range(ROW1, ROWN + 1):
    data = despesas[r - ROW1] if r - ROW1 < len(despesas) else [None] * 10
    fmts = [FMT_DATE, None, None, FMT_MONEY, None, None, None, FMT_DATE, None, None]
    aligns = [C_CENTER, C_LEFT, C_LEFT, C_RIGHT, C_LEFT, C_CENTER, C_CENTER,
              C_CENTER, C_CENTER, C_CENTER]
    for j in range(10):
        put(ws, r, 1 + j, data[j], fmt=fmts[j], align=aligns[j], unlock=True)
    # ranking de vencimentos p/ a aba Alertas (vencidas + próximas de 7 dias)
    f = (f'=IF(AND($I{r}="Não",$H{r}<>"",$H{r}<=TODAY()+7),'
         f'COUNTIFS($H${ROW1}:$H${ROWN},"<"&$H{r},$I${ROW1}:$I${ROWN},"Não",'
         f'$H${ROW1}:$H${ROWN},"<="&TODAY()+7)'
         f'+COUNTIFS($H${ROW1}:$H{r},$H{r},$I${ROW1}:$I{r},"Não"),"")')
    put(ws, r, 11, f, font=Font(color=CARD), align=C_CENTER)

for col, name in [("C", "CatDespesas"), ("E", "FormasPag"),
                  ("F", "ListaCartoes"), ("J", "ListaContas")]:
    dv = DataValidation(type="list", formula1=f"={name}", allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{col}{ROW1}:{col}{ROWN}")
dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"I{ROW1}:I{ROWN}")

ws.conditional_formatting.add(f"I{ROW1}:I{ROWN}", FormulaRule(
    formula=[f'$I{ROW1}="Sim"'], font=Font(color=GREEN_HI, bold=True)))
ws.conditional_formatting.add(f"I{ROW1}:I{ROWN}", FormulaRule(
    formula=[f'$I{ROW1}="Não"'], font=Font(color=RED, bold=True)))
ws.conditional_formatting.add(f"H{ROW1}:H{ROWN}", FormulaRule(
    formula=[f'AND($I{ROW1}="Não",$H{ROW1}<>"",$H{ROW1}<TODAY())'],
    fill=FILL_RED, font=Font(color=WHITE, bold=True)))
ws.conditional_formatting.add(f"H{ROW1}:H{ROWN}", FormulaRule(
    formula=[f'AND($I{ROW1}="Não",$H{ROW1}<>"",$H{ROW1}<=TODAY()+7)'],
    fill=FILL_AMBERD, font=Font(color=AMBER, bold=True)))

# ----------------------------------------------------------------------------
# 🏦 CONTAS
# ----------------------------------------------------------------------------
ws = ws_con
paint_bg(ws, 40, 12)
nav_bar(ws, SH_CON)
title(ws, 3, 2, "🏦 Contas Bancárias",
      "Entradas, saídas e transferências entram automaticamente no saldo.")

CON_H = ["Banco / Conta", "Saldo Inicial", "Entradas", "Saídas",
         "Transf. Recebidas", "Transf. Enviadas", "Saldo Atual"]
table_header(ws, 6, 1, CON_H)
set_widths(ws, [20, 15, 15, 15, 17, 17, 16])
CON_R1, CON_RN = 7, 14
saldos_ini = {"Itaú": 2500, "Nubank": 1200, "Caixa": 800, "Inter": 300,
              "Carteira": 150}
TR_R1, TR_RN = 22, 36
for i, r in enumerate(range(CON_R1, CON_RN + 1)):
    banco = CONTAS[i] if i < len(CONTAS) else None
    put(ws, r, 1, banco, unlock=True)
    put(ws, r, 2, saldos_ini.get(banco), fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 3, f'=IF($A{r}="","",SUMIFS({Q(SH_REC)}!$D:$D,{Q(SH_REC)}!$F:$F,$A{r},{Q(SH_REC)}!$G:$G,"Recebido"))',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=GREEN_HI))
    put(ws, r, 4, f'=IF($A{r}="","",SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$J:$J,$A{r},{Q(SH_DES)}!$I:$I,"Sim"))',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=RED))
    put(ws, r, 5, f'=IF($A{r}="","",SUMIFS($D${TR_R1}:$D${TR_RN},$C${TR_R1}:$C${TR_RN},$A{r}))',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 6, f'=IF($A{r}="","",SUMIFS($D${TR_R1}:$D${TR_RN},$B${TR_R1}:$B${TR_RN},$A{r}))',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 7, f'=IF($A{r}="","",$B{r}+$C{r}-$D{r}+$E{r}-$F{r})',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(bold=True, color=GREEN_HI))
TOT_R = CON_RN + 1
put(ws, TOT_R, 1, "TOTAL", font=F_HEAD, fill=FILL_HEADER)
for c in range(2, 8):
    L = get_column_letter(c)
    put(ws, TOT_R, c, f"=SUM({L}{CON_R1}:{L}{CON_RN})", fmt=FMT_MONEY,
        align=C_RIGHT, font=Font(bold=True, color=GREEN_HI), fill=FILL_HEADER)

section(ws, 19, 1, "🔁 Transferências entre contas")
table_header(ws, 21, 1, ["Data", "De (conta)", "Para (conta)", "Valor", "Observação"])
transfs = [(dt.date(ANO, 5, 10), "Nubank", "Carteira", 200, "Dinheiro p/ semana"),
           (dt.date(ANO, 6, 15), "Itaú", "Nubank", 400, "Cobrir fatura"),
           (dt.date(ANO, 7, 1), "Itaú", "Inter", 300, "Investimento")]
for i, r in enumerate(range(TR_R1, TR_RN + 1)):
    row = transfs[i] if i < len(transfs) else [None] * 5
    put(ws, r, 1, row[0], fmt=FMT_DATE, align=C_CENTER, unlock=True)
    put(ws, r, 2, row[1], unlock=True)
    put(ws, r, 3, row[2], unlock=True)
    put(ws, r, 4, row[3], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 5, row[4], unlock=True)
for col in ("B", "C"):
    dv = DataValidation(type="list", formula1="=ListaContas", allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"{col}{TR_R1}:{col}{TR_RN}")
ws.conditional_formatting.add(f"G{CON_R1}:G{CON_RN}", FormulaRule(
    formula=[f'AND($A{CON_R1}<>"",$G{CON_R1}<0)'], font=Font(color=RED, bold=True)))

# ----------------------------------------------------------------------------
# 💳 CARTÕES
# ----------------------------------------------------------------------------
ws = ws_car
paint_bg(ws, 32, 10)
nav_bar(ws, SH_CAR)
title(ws, 3, 2, "💳 Controle de Cartões",
      "A fatura soma as compras no crédito ainda não pagas na aba 💸 Despesas.")

CAR_H = ["Cartão", "Limite", "Fatura Atual", "Limite Disponível", "% Utilizado",
         "Uso do limite", "Dia Fechamento", "Dia Vencimento", "Melhor Dia de Compra"]
table_header(ws, 6, 1, CAR_H)
set_widths(ws, [15, 14, 14, 17, 12, 15, 15, 15, 19])
CAR_R1, CAR_RN = 7, 12
cartoes = [("Nubank", 4500, 28, 5), ("Itaú", 6000, 25, 3), ("Inter", 2000, 20, 27)]
for i, r in enumerate(range(CAR_R1, CAR_RN + 1)):
    row = cartoes[i] if i < len(cartoes) else (None, None, None, None)
    put(ws, r, 1, row[0], unlock=True)
    put(ws, r, 2, row[1], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 3, f'=IF($A{r}="","",SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$F:$F,$A{r},{Q(SH_DES)}!$E:$E,"Crédito",{Q(SH_DES)}!$I:$I,"Não"))',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=AMBER))
    put(ws, r, 4, f'=IF($A{r}="","",$B{r}-$C{r})', fmt=FMT_MONEY, align=C_RIGHT,
        font=Font(color=GREEN_HI, bold=True))
    put(ws, r, 5, f'=IF(OR($A{r}="",$B{r}=0),"",$C{r}/$B{r})', fmt=FMT_PCT,
        align=C_CENTER)
    put(ws, r, 6, f'=IF($E{r}="","",REPT("█",MIN(10,ROUND($E{r}*10,0)))&REPT("░",10-MIN(10,ROUND($E{r}*10,0))))',
        font=Font(color=GREEN, bold=True), align=C_CENTER)
    put(ws, r, 7, row[2], align=C_CENTER, unlock=True)
    put(ws, r, 8, row[3], align=C_CENTER, unlock=True)
    put(ws, r, 9, f'=IF($G{r}="","",IF($G{r}>=28,1,$G{r}+1))', align=C_CENTER,
        font=Font(color=GREEN_HI, bold=True))
ws.conditional_formatting.add(f"E{CAR_R1}:E{CAR_RN}", FormulaRule(
    formula=[f'AND($E{CAR_R1}<>"",$E{CAR_R1}>=0.8)'], font=Font(color=RED, bold=True)))
ws.conditional_formatting.add(f"E{CAR_R1}:E{CAR_RN}", FormulaRule(
    formula=[f'AND($E{CAR_R1}<>"",$E{CAR_R1}>=0.5)'], font=Font(color=AMBER, bold=True)))

section(ws, 15, 1, "📆 Parcelamentos em andamento")
PAR_H = ["Descrição", "Cartão", "Valor Total", "Nº Parcelas", "Valor da Parcela",
         "Parcelas Pagas", "Falta Pagar", "Situação"]
table_header(ws, 17, 1, PAR_H)
PAR_R1, PAR_RN = 18, 27
parcelamentos = [("Notebook Dell", "Itaú", 3600, 12, 6),
                 ("Smartphone", "Nubank", 2400, 10, 3)]
for i, r in enumerate(range(PAR_R1, PAR_RN + 1)):
    row = parcelamentos[i] if i < len(parcelamentos) else (None,) * 5
    put(ws, r, 1, row[0], unlock=True)
    put(ws, r, 2, row[1], unlock=True)
    put(ws, r, 3, row[2], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 4, row[3], align=C_CENTER, unlock=True)
    put(ws, r, 5, f'=IF(OR($C{r}="",$D{r}=""),"",$C{r}/$D{r})', fmt=FMT_MONEY,
        align=C_RIGHT)
    put(ws, r, 6, row[4], align=C_CENTER, unlock=True)
    put(ws, r, 7, f'=IF($E{r}="","",$C{r}-$E{r}*$F{r})', fmt=FMT_MONEY,
        align=C_RIGHT, font=Font(color=AMBER))
    put(ws, r, 8, f'=IF($E{r}="","",IF($F{r}>=$D{r},"✅ Quitado","⏳ "&$F{r}&"/"&$D{r}&" pagas"))',
        align=C_CENTER)
dv = DataValidation(type="list", formula1="=ListaCartoes", allow_blank=True)
ws.add_data_validation(dv); dv.add(f"B{PAR_R1}:B{PAR_RN}")
dv = DataValidation(type="list", formula1="=ListaCartoes", allow_blank=True)
ws.add_data_validation(dv); dv.add(f"A{CAR_R1}:A{CAR_RN}")

# ----------------------------------------------------------------------------
# 🎯 METAS
# ----------------------------------------------------------------------------
ws = ws_met
paint_bg(ws, 24, 12)
nav_bar(ws, SH_MET)
title(ws, 3, 2, "🎯 Metas Financeiras",
      "Informe a meta, o quanto já guardou e o aporte mensal — o resto é automático.")

MET_H = ["Meta", "Valor Desejado", "Valor Guardado", "Valor Restante",
         "Progresso", "%", "Aporte Mensal", "Meses Restantes",
         "Previsão de Conclusão", "Status"]
table_header(ws, 6, 1, MET_H)
set_widths(ws, [26, 15, 15, 15, 15, 9, 14, 10, 15, 18])
MET_R1, MET_RN = 7, 14
metas = [("Reserva de emergência", 15000, 8200, 600),
         ("Viagem de férias", 6000, 2100, 400),
         ("Notebook novo", 4500, 4500, 0),
         ("Entrada do apartamento", 40000, 5500, 800)]
for i, r in enumerate(range(MET_R1, MET_RN + 1)):
    row = metas[i] if i < len(metas) else (None,) * 4
    put(ws, r, 1, row[0], unlock=True)
    put(ws, r, 2, row[1], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 3, row[2], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 4, f'=IF($B{r}="","",MAX($B{r}-$C{r},0))', fmt=FMT_MONEY,
        align=C_RIGHT)
    put(ws, r, 5, f'=IF($B{r}="","",REPT("█",MIN(10,ROUND($F{r}*10,0)))&REPT("░",10-MIN(10,ROUND($F{r}*10,0))))',
        font=Font(color=GREEN, bold=True), align=C_CENTER)
    put(ws, r, 6, f'=IF($B{r}="","",MIN($C{r}/$B{r},1))', fmt='0%', align=C_CENTER)
    put(ws, r, 7, row[3], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 8, f'=IF(OR($B{r}="",$C{r}>=$B{r}),"",IF($G{r}>0,ROUNDUP(($B{r}-$C{r})/$G{r},0),"—"))',
        align=C_CENTER)
    put(ws, r, 9, f'=IF(OR($B{r}="",$C{r}>=$B{r},$G{r}<=0),"",EDATE(TODAY(),ROUNDUP(($B{r}-$C{r})/$G{r},0)))',
        fmt=FMT_MESANO, align=C_CENTER)
    put(ws, r, 10, f'=IF($B{r}="","",IF($C{r}>=$B{r},"🎉 Atingida!",IF($F{r}>=0.7,"🔥 Quase lá",IF($F{r}>=0.3,"💪 Em progresso","🚀 Começando"))))',
        align=C_CENTER)
ws.conditional_formatting.add(f"J{MET_R1}:J{MET_RN}", FormulaRule(
    formula=[f'AND($B{MET_R1}<>"",$C{MET_R1}>=$B{MET_R1})'],
    fill=FILL_GREEND, font=Font(color=GREEN_HI, bold=True)))

# ----------------------------------------------------------------------------
# 📅 PLANEJAMENTO MENSAL
# ----------------------------------------------------------------------------
ws = ws_pla
paint_bg(ws, 26, 10)
nav_bar(ws, SH_PLA)
title(ws, 3, 2, "📅 Planejamento Mensal")
put(ws, 4, 2, f'="Orçamento de "&INDEX({Q(SH_CFG)}!$H$7:$H$18,{MES_CELL})&" de "&{ANO_CELL}&" — mude o mês no 📊 Dashboard"',
    font=F_SUB, fill=FILL_BG, border=None)

PLA_H = ["Categoria", "Orçamento", "Valor Gasto", "Diferença", "% Usado",
         "Consumo", "Situação"]
table_header(ws, 6, 1, PLA_H)
set_widths(ws, [18, 13, 13, 13, 10, 15, 15])
PLA_R1 = 7
orcamentos = {"Moradia": 1900, "Alimentação": 900, "Transporte": 350,
              "Saúde": 250, "Educação": 150, "Lazer": 250, "Vestuário": 150,
              "Assinaturas": 120, "Contas Fixas": 400, "Investimentos": 600,
              "Outros": 700}
PLA_RN = PLA_R1 + len(CAT_DES) - 1
for i, cat in enumerate(CAT_DES):
    r = PLA_R1 + i
    put(ws, r, 1, cat)
    put(ws, r, 2, orcamentos[cat], fmt=FMT_MONEY, align=C_RIGHT, unlock=True)
    put(ws, r, 3, f'=SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$C:$C,$A{r},{Q(SH_DES)}!$A:$A,">="&{INI_MES},{Q(SH_DES)}!$A:$A,"<="&{FIM_MES})',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 4, f'=$B{r}-$C{r}', fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 5, f'=IFERROR($C{r}/$B{r},0)', fmt='0%', align=C_CENTER)
    put(ws, r, 6, f'=REPT("█",MIN(10,ROUND($E{r}*10,0)))&REPT("░",10-MIN(10,ROUND($E{r}*10,0)))',
        font=Font(color=GREEN, bold=True), align=C_CENTER)
    put(ws, r, 7, f'=IF($C{r}>$B{r},"🔴 Estourou",IF($E{r}>=0.8,"🟡 Atenção","🟢 Dentro"))',
        align=C_CENTER)
TOTP = PLA_RN + 1
put(ws, TOTP, 1, "TOTAL", font=F_HEAD, fill=FILL_HEADER)
for c, fmt in [(2, FMT_MONEY), (3, FMT_MONEY), (4, FMT_MONEY)]:
    L = get_column_letter(c)
    put(ws, TOTP, c, f"=SUM({L}{PLA_R1}:{L}{PLA_RN})", fmt=fmt, align=C_RIGHT,
        font=Font(bold=True, color=GREEN_HI), fill=FILL_HEADER)
put(ws, TOTP, 5, f'=IFERROR($C{TOTP}/$B{TOTP},0)', fmt='0%', align=C_CENTER,
    font=Font(bold=True, color=GREEN_HI), fill=FILL_HEADER)

rng = f"G{PLA_R1}:G{PLA_RN}"
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'$C{PLA_R1}>$B{PLA_R1}'], fill=FILL_RED, font=Font(color=WHITE, bold=True)))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'$E{PLA_R1}>=0.8'], fill=FILL_AMBERD, font=Font(color=AMBER, bold=True)))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'$E{PLA_R1}<0.8'], fill=FILL_GREEND, font=Font(color=GREEN_HI, bold=True)))

# ----------------------------------------------------------------------------
# 📈 RELATÓRIOS
# ----------------------------------------------------------------------------
ws = ws_rel
paint_bg(ws, 60, 14)
nav_bar(ws, SH_REL)
title(ws, 3, 2, "📈 Relatórios")
put(ws, 4, 2, f'="Ano de referência: "&{ANO_CELL}&" (definido no 📊 Dashboard)"',
    font=F_SUB, fill=FILL_BG, border=None)

# --- Resumo mensal (A6:E19) -------------------------------------------------
section(ws, 6, 1, "Resumo mensal")
table_header(ws, 7, 1, ["Mês", "Receitas", "Despesas", "Saldo", "Patrimônio"])
set_widths(ws, [9, 14, 14, 14, 15, 3, 18, 13, 9, 3, 14, 14, 14, 14])
REL_R1 = 8
for i, mes in enumerate(MESES):
    r = REL_R1 + i
    m = i + 1
    ini = f"DATE({ANO_CELL},{m},1)"
    fim = f"EOMONTH(DATE({ANO_CELL},{m},1),0)"
    put(ws, r, 1, mes, align=C_CENTER, font=F_HEAD)
    put(ws, r, 2, f'=SUMIFS({Q(SH_REC)}!$D:$D,{Q(SH_REC)}!$A:$A,">="&{ini},{Q(SH_REC)}!$A:$A,"<="&{fim})',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=GREEN_HI))
    put(ws, r, 3, f'=SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$A:$A,">="&{ini},{Q(SH_DES)}!$A:$A,"<="&{fim})',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 4, f'=$B{r}-$C{r}', fmt=FMT_MONEY, align=C_RIGHT)
    prev = f"$E{r-1}" if i else f"{Q(SH_CON)}!$B${TOT_R}"
    put(ws, r, 5, f'={prev}+$D{r}', fmt=FMT_MONEY, align=C_RIGHT,
        font=Font(bold=True, color=GREEN_HI))
REL_RN = REL_R1 + 11
put(ws, REL_RN + 1, 1, "TOTAL", font=F_HEAD, fill=FILL_HEADER, align=C_CENTER)
for c in (2, 3, 4):
    L = get_column_letter(c)
    put(ws, REL_RN + 1, c, f"=SUM({L}{REL_R1}:{L}{REL_RN})", fmt=FMT_MONEY,
        align=C_RIGHT, font=Font(bold=True, color=GREEN_HI), fill=FILL_HEADER)
put(ws, REL_RN + 1, 5, "", fill=FILL_HEADER)

# --- Gastos por categoria no mês selecionado (G6:I19) ------------------------
section(ws, 6, 7, "Gastos por categoria (mês selecionado)")
table_header(ws, 7, 7, ["Categoria", "Valor", "%"])
GC_R1 = 8
for i, cat in enumerate(CAT_DES):
    r = GC_R1 + i
    put(ws, r, 7, cat)
    put(ws, r, 8, f'=SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$C:$C,$G{r},{Q(SH_DES)}!$A:$A,">="&{INI_MES},{Q(SH_DES)}!$A:$A,"<="&{FIM_MES})',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 9, f'=IFERROR($H{r}/SUM($H${GC_R1}:$H${GC_R1+10}),0)', fmt='0%',
        align=C_CENTER)
GC_RN = GC_R1 + 10

# --- Ranking (A23:D33) --------------------------------------------------------
section(ws, 22, 1, "🏆 Ranking — 10 maiores despesas lançadas")
table_header(ws, 23, 1, ["Nº", "Descrição", "Categoria", "Valor"])
ws.merge_cells("B23:C23")
RK_R1 = 24
for k in range(1, 11):
    r = RK_R1 + k - 1
    put(ws, r, 1, k, align=C_CENTER, font=F_HEAD)
    put(ws, r, 2, f'=IFERROR(INDEX({Q(SH_DES)}!$B${ROW1}:$B${ROWN},MATCH($D{r},{Q(SH_DES)}!$D${ROW1}:$D${ROWN},0)),"")')
    ws.merge_cells(f"B{r}:C{r}")
    put(ws, r, 4, f'=IFERROR(LARGE({Q(SH_DES)}!$D${ROW1}:$D${ROWN},{k}),"")',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=AMBER))

# --- Comparativo anual (G23:J27) ----------------------------------------------
section(ws, 22, 7, "📅 Comparativo anual")
table_header(ws, 23, 7, ["Ano", "Receitas", "Despesas"])
for i, expr in enumerate([f"{ANO_CELL}-1", f"{ANO_CELL}"]):
    r = 24 + i
    put(ws, r, 7, f"={expr}", align=C_CENTER, font=F_HEAD)
    put(ws, r, 8, f'=SUMIFS({Q(SH_REC)}!$D:$D,{Q(SH_REC)}!$A:$A,">="&DATE({expr},1,1),{Q(SH_REC)}!$A:$A,"<="&DATE({expr},12,31))',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=GREEN_HI))
    put(ws, r, 9, f'=SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$A:$A,">="&DATE({expr},1,1),{Q(SH_DES)}!$A:$A,"<="&DATE({expr},12,31))',
        fmt=FMT_MONEY, align=C_RIGHT)

# --- Gráfico de barras por categoria na própria aba ---------------------------
bar2 = BarChart()
bar2.type = "bar"
bar2.title = "Gastos por categoria (mês)"
bar2.height, bar2.width = 10, 16
data = Reference(ws, min_col=8, min_row=7, max_row=GC_RN)
cats = Reference(ws, min_col=7, min_row=GC_R1, max_row=GC_RN)
bar2.add_data(data, titles_from_data=True)
bar2.set_categories(cats)
bar2.series[0].graphicalProperties.solidFill = GREEN
bar2.legend = None
ws.add_chart(bar2, "G36")

# ----------------------------------------------------------------------------
# 🔔 ALERTAS
# ----------------------------------------------------------------------------
ws = ws_ale
paint_bg(ws, 60, 10)
nav_bar(ws, SH_ALE)
title(ws, 3, 2, "🔔 Alertas Inteligentes",
      "Tudo aqui é automático — os avisos aparecem conforme seus lançamentos.")

section(ws, 6, 1, "Resumo")
resumo = [
    ("🚨 Orçamentos estourados no mês",
     f'=SUMPRODUCT(({Q(SH_PLA)}!$C${PLA_R1}:$C${PLA_RN}>{Q(SH_PLA)}!$B${PLA_R1}:$B${PLA_RN})*1)'),
    ("⏰ Contas a vencer nos próximos 7 dias",
     f'=COUNTIFS({Q(SH_DES)}!$I:$I,"Não",{Q(SH_DES)}!$H:$H,">="&TODAY(),{Q(SH_DES)}!$H:$H,"<="&TODAY()+7)'),
    ("🔴 Contas vencidas e não pagas",
     f'=COUNTIFS({Q(SH_DES)}!$I:$I,"Não",{Q(SH_DES)}!$H:$H,"<"&TODAY())'),
    ("🎉 Metas atingidas",
     f'=SUMPRODUCT(({Q(SH_MET)}!$C${MET_R1}:$C${MET_RN}>={Q(SH_MET)}!$B${MET_R1}:$B${MET_RN})*({Q(SH_MET)}!$B${MET_R1}:$B${MET_RN}>0))'),
]
set_widths(ws, [4, 14, 36, 14, 10, 14])
ALE_RES_R1 = 7
for i, (label, f) in enumerate(resumo):
    r = ALE_RES_R1 + i
    put(ws, r, 2, label)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    put(ws, r, 5, f, align=C_CENTER, font=Font(size=13, bold=True, color=GREEN_HI))
    put(ws, r, 6, f'=IF($E{r}=0,"✅ OK","⚠️ Verificar")', align=C_CENTER)
ws.conditional_formatting.add(f"E{ALE_RES_R1}:F{ALE_RES_R1+2}", FormulaRule(
    formula=[f'$E{ALE_RES_R1}>0'], font=Font(color=RED, bold=True)))

section(ws, 13, 1, "⏰ Vencimentos próximos (inclui vencidas)")
table_header(ws, 14, 1, ["Nº", "Vencimento", "Descrição", "Valor", "Dias", "Situação"])
VC_R1 = 15
for k in range(1, 11):
    r = VC_R1 + k - 1
    put(ws, r, 1, k, align=C_CENTER, font=F_HEAD)
    put(ws, r, 2, f'=IFERROR(INDEX({Q(SH_DES)}!$H${ROW1}:$H${ROWN},MATCH({k},{Q(SH_DES)}!$K${ROW1}:$K${ROWN},0)),"")',
        fmt=FMT_DATE, align=C_CENTER)
    put(ws, r, 3, f'=IFERROR(INDEX({Q(SH_DES)}!$B${ROW1}:$B${ROWN},MATCH({k},{Q(SH_DES)}!$K${ROW1}:$K${ROWN},0)),"")')
    put(ws, r, 4, f'=IFERROR(INDEX({Q(SH_DES)}!$D${ROW1}:$D${ROWN},MATCH({k},{Q(SH_DES)}!$K${ROW1}:$K${ROWN},0)),"")',
        fmt=FMT_MONEY, align=C_RIGHT)
    put(ws, r, 5, f'=IF($B{r}="","",$B{r}-TODAY())', fmt='0', align=C_CENTER)
    put(ws, r, 6, f'=IF($B{r}="","",IF($E{r}<0,"🔴 VENCIDA",IF($E{r}<=3,"🟠 Urgente","🟡 Em breve")))',
        align=C_CENTER)
ws.conditional_formatting.add(f"B{VC_R1}:F{VC_R1+9}", FormulaRule(
    formula=[f'AND($B{VC_R1}<>"",$E{VC_R1}<0)'], fill=FILL_RED))
ws.conditional_formatting.add(f"B{VC_R1}:F{VC_R1+9}", FormulaRule(
    formula=[f'AND($B{VC_R1}<>"",$E{VC_R1}<=3)'], fill=FILL_AMBERD))

section(ws, 27, 1, "🚨 Orçamentos estourados (mês selecionado)")
table_header(ws, 28, 1, ["", "Categoria", "Estouro (R$)", "", "", ""])
EO_R1 = 29
for i in range(len(CAT_DES)):
    r = EO_R1 + i
    pr = PLA_R1 + i
    put(ws, r, 2, f'=IF({Q(SH_PLA)}!$C${pr}>{Q(SH_PLA)}!$B${pr},{Q(SH_PLA)}!$A${pr},"")')
    put(ws, r, 3, f'=IF({Q(SH_PLA)}!$C${pr}>{Q(SH_PLA)}!$B${pr},{Q(SH_PLA)}!$C${pr}-{Q(SH_PLA)}!$B${pr},"")',
        fmt=FMT_MONEY, align=C_RIGHT, font=Font(color=RED, bold=True))

section(ws, 42, 1, "🎉 Metas atingidas")
GM_R1 = 43
for i in range(MET_RN - MET_R1 + 1):
    r = GM_R1 + i
    mr = MET_R1 + i
    put(ws, r, 2, f'=IF(AND({Q(SH_MET)}!$B${mr}>0,{Q(SH_MET)}!$C${mr}>={Q(SH_MET)}!$B${mr}),"🎉 Parabéns! Meta """&{Q(SH_MET)}!$A${mr}&""" atingida!","")',
        font=Font(color=GREEN_HI, bold=True))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

# ----------------------------------------------------------------------------
# 📊 DASHBOARD
# ----------------------------------------------------------------------------
ws = ws_dash
paint_bg(ws, 56, 14)
nav_bar(ws, SH_DASH)
set_widths(ws, [2] + [14] * 12)

put(ws, 3, 2, "💚 PAINEL FINANCEIRO PESSOAL", font=F_TITLE, fill=FILL_BG, border=None)
ws.merge_cells("B3:H3")
put(ws, 3, 9, "Mês:", font=F_LABEL, fill=FILL_BG, align=C_RIGHT, border=None)
put(ws, 3, 10, 7, font=Font(size=12, bold=True, color=GREEN_HI),
    fill=FILL_CARD, align=C_CENTER, unlock=True)
put(ws, 3, 11, "Ano:", font=F_LABEL, fill=FILL_BG, align=C_RIGHT, border=None)
put(ws, 3, 12, ANO, font=Font(size=12, bold=True, color=GREEN_HI),
    fill=FILL_CARD, align=C_CENTER, unlock=True)
dv = DataValidation(type="list", formula1="=ListaMeses", allow_blank=False)
ws.add_data_validation(dv); dv.add("J3")
dv = DataValidation(type="list", formula1="=ListaAnos", allow_blank=False)
ws.add_data_validation(dv); dv.add("L3")
put(ws, 4, 2, f'="Visão de "&INDEX({Q(SH_CFG)}!$H$7:$H$18,$J$3)&" de "&$L$3',
    font=F_SUB, fill=FILL_BG, border=None)

REC_MES = f'SUMIFS({Q(SH_REC)}!$D:$D,{Q(SH_REC)}!$A:$A,">="&{INI_MES},{Q(SH_REC)}!$A:$A,"<="&{FIM_MES})'
DES_MES = f'SUMIFS({Q(SH_DES)}!$D:$D,{Q(SH_DES)}!$A:$A,">="&{INI_MES},{Q(SH_DES)}!$A:$A,"<="&{FIM_MES})'

cards = [
    ("💼 SALDO ATUAL",      f"={Q(SH_CON)}!$G${TOT_R}", FMT_MONEY,
     '="Somando todas as contas"'),
    ("💰 RECEITAS DO MÊS",  f"={REC_MES}", FMT_MONEY,
     '="Entradas do período"'),
    ("💸 DESPESAS DO MÊS",  f"={DES_MES}", FMT_MONEY,
     '="Saídas do período"'),
    ("🐷 ECONOMIA DO MÊS",  f"={REC_MES}-{DES_MES}", FMT_MONEY,
     f'=IF({REC_MES}=0,"","Taxa de poupança: "&TEXT(({REC_MES}-{DES_MES})/{REC_MES},"0%"))'),
    ("📊 % DE GASTOS",      f"=IFERROR({DES_MES}/{REC_MES},0)", FMT_PCT,
     f'=IF($J$7>0.9,"🚨 Gastos muito altos",IF($J$7>0.7,"⚠️ Atenção aos gastos","✅ Sob controle"))'),
]
for i, (label, formula, fmt, sub) in enumerate(cards):
    c0 = 2 + i * 2   # B, D, F, H, J
    L0, L1 = get_column_letter(c0), get_column_letter(c0 + 1)
    ws.merge_cells(f"{L0}6:{L1}6")
    ws.merge_cells(f"{L0}7:{L1}7")
    ws.merge_cells(f"{L0}8:{L1}8")
    put(ws, 6, c0, label, font=F_LABEL, fill=FILL_CARD, align=C_CENTER, border=None)
    put(ws, 6, c0 + 1, None, fill=FILL_CARD, border=None)
    put(ws, 7, c0, formula, font=F_KPI_G, fill=FILL_CARD, fmt=fmt,
        align=C_CENTER, border=None)
    put(ws, 7, c0 + 1, None, fill=FILL_CARD, border=None)
    put(ws, 8, c0, sub, font=F_SUB, fill=FILL_CARD, align=C_CENTER, border=None)
    c = put(ws, 8, c0 + 1, None, fill=FILL_CARD, border=None)
    for col in (c0, c0 + 1):
        ws.cell(row=8, column=col).border = ACCENT_BOT
ws.row_dimensions[6].height = 18
ws.row_dimensions[7].height = 30
ws.row_dimensions[8].height = 18
# despesas e % em branco/amarelo p/ contraste com receitas verdes
ws.cell(row=7, column=6).font = F_KPI
ws.cell(row=7, column=10).font = F_KPI

ws.conditional_formatting.add("H7", FormulaRule(
    formula=['$H$7<0'], font=Font(size=16, bold=True, color=RED)))
ws.conditional_formatting.add("J7", FormulaRule(
    formula=['$J$7>0.9'], font=Font(size=16, bold=True, color=RED)))
ws.conditional_formatting.add("J7", FormulaRule(
    formula=['$J$7>0.7'], font=Font(size=16, bold=True, color=AMBER)))

# Banner de situação
ws.merge_cells("B10:K10")
put(ws, 10, 2,
    f'=IF({REC_MES}=0,"ℹ️ Nenhuma receita registrada neste mês.",'
    f'IF({REC_MES}-{DES_MES}<0,"🔴 CRÍTICO: você gastou mais do que ganhou neste mês!",'
    f'IF({REC_MES}-{DES_MES}<0.2*{REC_MES},"🟡 ATENÇÃO: economia abaixo de 20% da renda.",'
    f'"🟢 EXCELENTE: você está economizando mais de 20% da renda!")))',
    font=Font(size=12, bold=True, color=WHITE), fill=FILL_CARD, align=C_CENTER,
    border=None)
ws.row_dimensions[10].height = 26
ws.conditional_formatting.add("B10", FormulaRule(
    formula=['ISNUMBER(SEARCH("CRÍTICO",$B$10))'], fill=FILL_RED))
ws.conditional_formatting.add("B10", FormulaRule(
    formula=['ISNUMBER(SEARCH("ATENÇÃO",$B$10))'], fill=FILL_AMBERD))
ws.conditional_formatting.add("B10", FormulaRule(
    formula=['ISNUMBER(SEARCH("EXCELENTE",$B$10))'], fill=FILL_GREEND))

# Mini-alertas
alertas_mini = [
    ("🚨 Orçamentos estourados", f"={Q(SH_ALE)}!$E${ALE_RES_R1}"),
    ("⏰ Contas a vencer (7 dias)", f"={Q(SH_ALE)}!$E${ALE_RES_R1+1}"),
    ("🔴 Contas vencidas", f"={Q(SH_ALE)}!$E${ALE_RES_R1+2}"),
    ("🎯 Metas atingidas", f"={Q(SH_ALE)}!$E${ALE_RES_R1+3}"),
]
for i, (label, f) in enumerate(alertas_mini):
    c0 = 2 + i * 3
    L0, L1 = get_column_letter(c0), get_column_letter(c0 + 1)
    ws.merge_cells(f"{L0}12:{L1}12")
    put(ws, 12, c0, label, font=F_LABEL, fill=FILL_CARD, align=C_CENTER, border=None)
    put(ws, 12, c0 + 1, None, fill=FILL_CARD, border=None)
    put(ws, 12, c0 + 2, f, font=Font(size=13, bold=True, color=GREEN_HI),
        fill=FILL_CARD, align=C_CENTER, border=None)
ws.row_dimensions[12].height = 24
put(ws, 13, 2, '="👉 Detalhes na aba 🔔 Alertas"', font=F_SUB, fill=FILL_BG,
    border=None)

# --- Gráficos ---------------------------------------------------------------
GREENS = ["00C853", "69F0AE", "2E7D32", "A5D6A7", "00E676", "1B5E20",
          "66BB6A", "81C784", "4CAF50", "9E9E9E", "616161"]

pie = PieChart()
pie.title = "Gastos por categoria (mês)"
pie.height, pie.width = 9.5, 11
data = Reference(ws_rel, min_col=8, min_row=7, max_row=GC_RN)
cats = Reference(ws_rel, min_col=7, min_row=GC_R1, max_row=GC_RN)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
for idx, color in enumerate(GREENS):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)
ws.add_chart(pie, "B15")

bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.title = "Receitas × Despesas por mês"
bar.height, bar.width = 9.5, 15
data = Reference(ws_rel, min_col=2, max_col=3, min_row=7, max_row=REL_RN)
cats = Reference(ws_rel, min_col=1, min_row=REL_R1, max_row=REL_RN)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = GREEN
bar.series[1].graphicalProperties.solidFill = "455A64"
bar.legend.position = "b"
ws.add_chart(bar, "G15")

line = LineChart()
line.title = "Evolução do patrimônio no ano"
line.height, line.width = 9, 24
data = Reference(ws_rel, min_col=5, min_row=7, max_row=REL_RN)
cats = Reference(ws_rel, min_col=1, min_row=REL_R1, max_row=REL_RN)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
s = line.series[0]
s.smooth = True
s.graphicalProperties.line.solidFill = GREEN
s.graphicalProperties.line.width = 28575  # 2.25 pt
line.legend = None
ws.add_chart(line, "B36")

# ----------------------------------------------------------------------------
# Proteção (sem senha) — células de entrada ficam desbloqueadas
# ----------------------------------------------------------------------------
for ws in (ws_dash, ws_rec, ws_des, ws_con, ws_car, ws_met, ws_pla,
           ws_rel, ws_ale):
    protect(ws)

wb.active = 0
OUT = "Planilha_Financeira_Pessoal.xlsx"
wb.save(OUT)
print(f"OK -> {OUT}")
